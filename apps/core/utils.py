"""
Utilidades generales para el proyecto SIGAP.
"""
import os
import uuid
from datetime import datetime, date
from decimal import Decimal
from django.utils import timezone
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from PIL import Image
from io import BytesIO


def generate_unique_code(prefix='', length=8):
    """
    Genera un código único con prefijo opcional.
    
    Args:
        prefix: Prefijo para el código
        length: Longitud del código (sin contar el prefijo)
    
    Returns:
        str: Código único generado
    """
    unique_id = uuid.uuid4().hex[:length].upper()
    return f"{prefix}{unique_id}" if prefix else unique_id


def resize_image(image_file, max_width=1200, max_height=1200, quality=85):
    """
    Redimensiona una imagen manteniendo la proporción.
    
    Args:
        image_file: Archivo de imagen
        max_width: Ancho máximo
        max_height: Altura máxima
        quality: Calidad de compresión (1-100)
    
    Returns:
        BytesIO: Imagen redimensionada
    """
    img = Image.open(image_file)
    
    # Convertir a RGB si es necesario
    if img.mode in ('RGBA', 'LA', 'P'):
        background = Image.new('RGB', img.size, (255, 255, 255))
        background.paste(img, mask=img.split()[-1] if img.mode != 'P' else None)
        img = background
    
    # Calcular nuevas dimensiones manteniendo proporción
    ratio = min(max_width / img.width, max_height / img.height)
    if ratio < 1:
        new_size = (int(img.width * ratio), int(img.height * ratio))
        img = img.resize(new_size, Image.Resampling.LANCZOS)
    
    # Guardar en buffer
    output = BytesIO()
    img.save(output, format='JPEG', quality=quality, optimize=True)
    output.seek(0)
    
    return output


def save_resized_image(image_field, upload_path, max_width=1200, max_height=1200):
    """
    Guarda una imagen redimensionada.
    
    Args:
        image_field: Campo de imagen del modelo
        upload_path: Ruta de subida
        max_width: Ancho máximo
        max_height: Altura máxima
    
    Returns:
        str: Ruta del archivo guardado
    """
    if not image_field:
        return None
    
    resized = resize_image(image_field, max_width, max_height)
    filename = f"{uuid.uuid4().hex}.jpg"
    filepath = os.path.join(upload_path, filename)
    
    saved_path = default_storage.save(filepath, ContentFile(resized.read()))
    return saved_path


def format_currency(value, currency='MXN', decimals=2):
    """
    Formatea un valor como moneda.
    
    Args:
        value: Valor numérico
        currency: Código de moneda
        decimals: Número de decimales
    
    Returns:
        str: Valor formateado
    """
    if value is None:
        return '-'
    
    symbols = {
        'MXN': '$',
        'USD': 'US$',
        'EUR': '€',
    }
    
    symbol = symbols.get(currency, currency)
    return f"{symbol}{value:,.{decimals}f}"


def get_month_name(month_number):
    """
    Retorna el nombre del mes en español.
    
    Args:
        month_number: Número del mes (1-12)
    
    Returns:
        str: Nombre del mes
    """
    months = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    return months[month_number - 1] if 1 <= month_number <= 12 else ''


def calculate_depreciation(initial_value, years, current_year):
    """
    Calcula la depreciación lineal de un activo.
    
    Args:
        initial_value: Valor inicial
        years: Años de vida útil
        current_year: Año actual de depreciación
    
    Returns:
        dict: Información de depreciación
    """
    if not initial_value or not years or years <= 0:
        return None
    
    annual_depreciation = initial_value / years
    accumulated_depreciation = min(annual_depreciation * current_year, initial_value)
    current_value = max(initial_value - accumulated_depreciation, 0)
    
    return {
        'annual_depreciation': round(annual_depreciation, 2),
        'accumulated_depreciation': round(accumulated_depreciation, 2),
        'current_value': round(current_value, 2),
        'depreciation_percentage': round((accumulated_depreciation / initial_value) * 100, 2) if initial_value else 0
    }


def get_file_extension(filename):
    """
    Obtiene la extensión de un archivo.
    
    Args:
        filename: Nombre del archivo
    
    Returns:
        str: Extensión en minúsculas
    """
    return os.path.splitext(filename)[1].lower()


def validate_file_type(filename, allowed_extensions):
    """
    Valida si un archivo tiene una extensión permitida.
    
    Args:
        filename: Nombre del archivo
        allowed_extensions: Lista de extensiones permitidas
    
    Returns:
        bool: True si es válido
    """
    ext = get_file_extension(filename)
    return ext in allowed_extensions


def truncate_text(text, max_length=100, suffix='...'):
    """
    Trunca un texto a una longitud máxima.
    
    Args:
        text: Texto a truncar
        max_length: Longitud máxima
        suffix: Sufijo para indicar truncamiento
    
    Returns:
        str: Texto truncado
    """
    if not text:
        return ''
    
    if len(text) <= max_length:
        return text
    
    return text[:max_length - len(suffix)] + suffix


def get_client_ip(request):
    """
    Obtiene la IP real del cliente.
    
    Args:
        request: HttpRequest
    
    Returns:
        str: Dirección IP
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def days_between(date1, date2):
    """
    Calcula los días entre dos fechas.
    
    Args:
        date1: Primera fecha
        date2: Segunda fecha
    
    Returns:
        int: Número de días
    """
    if isinstance(date1, datetime):
        date1 = date1.date()
    if isinstance(date2, datetime):
        date2 = date2.date()
    
    return (date2 - date1).days


def is_date_due(check_date, days_warning=30):
    """
    Verifica si una fecha está próxima o vencida.
    
    Args:
        check_date: Fecha a verificar
        days_warning: Días de anticipación para advertencia
    
    Returns:
        dict: Estado de la fecha
    """
    if not check_date:
        return {'status': 'unknown', 'days': None}
    
    if isinstance(check_date, datetime):
        check_date = check_date.date()
    
    today = date.today()
    days = (check_date - today).days
    
    if days < 0:
        return {'status': 'overdue', 'days': days}
    elif days <= days_warning:
        return {'status': 'warning', 'days': days}
    else:
        return {'status': 'ok', 'days': days}


class ExportMixin:
    """
    Mixin para exportar datos a Excel.
    """
    
    def export_to_excel(self, queryset, filename, columns):
        """
        Exporta un queryset a Excel.
        
        Args:
            queryset: QuerySet a exportar
            filename: Nombre del archivo
            columns: Diccionario {campo: título}
        
        Returns:
            HttpResponse: Archivo Excel
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        
        for col, (field, title) in enumerate(columns.items(), 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, obj in enumerate(queryset, 2):
            for col, field in enumerate(columns.keys(), 1):
                value = getattr(obj, field, '')
                if callable(value):
                    value = value()
                ws.cell(row=row, column=col, value=str(value) if value else '')
        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        wb.save(response)
        return response
