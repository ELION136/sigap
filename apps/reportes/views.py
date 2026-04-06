"""
Vistas para la app reportes del proyecto SIGAP.
"""
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import TemplateView
from django.db.models import Count, Q, Sum, Avg
from django.utils import timezone
from datetime import timedelta, datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.activos.models import Activo, TipoActivo, Estado, EquipoFuncional
from apps.mantenimiento.models import Mantenimiento
from apps.organizacion.models import Planta, Area, Responsable


class ReporteIndexView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Vista principal de reportes."""
    template_name = 'reportes/index.html'
    permission_required = 'reportes.view_reporte'


@login_required
@permission_required('reportes.view_reporte')
def reporte_activos(request):
    """Genera reporte de activos con filtros."""
    # Filtros
    planta_id = request.GET.get('planta')
    tipo_id = request.GET.get('tipo')
    estado_id = request.GET.get('estado')
    responsable_id = request.GET.get('responsable')
    
    queryset = Activo.objects.select_related(
        'tipo', 'estado', 'tipo_propiedad', 'responsable', 'ubicacion'
    )
    
    if planta_id:
        queryset = queryset.filter(ubicacion__planta_id=planta_id)
    if tipo_id:
        queryset = queryset.filter(tipo_id=tipo_id)
    if estado_id:
        queryset = queryset.filter(estado_id=estado_id)
    if responsable_id:
        queryset = queryset.filter(responsable_id=responsable_id)
    
    # Estadísticas
    stats = {
        'total': queryset.count(),
        'por_tipo': queryset.values('tipo__nombre').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad'),
        'por_estado': queryset.values('estado__nombre', 'estado__color').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad'),
        'por_planta': queryset.values('ubicacion__planta__nombre').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad'),
        'valor_total': queryset.aggregate(
            total=Sum('valor_actual')
        )['total'] or 0,
    }
    
    context = {
        'activos': queryset[:100],  # Limitar a 100 para la vista
        'stats': stats,
        'plantas': Planta.objects.filter(activa=True),
        'tipos': TipoActivo.objects.filter(activo=True),
        'estados': Estado.objects.filter(activo=True),
        'responsables': Responsable.objects.filter(activo=True),
        'filtros': {
            'planta': planta_id,
            'tipo': tipo_id,
            'estado': estado_id,
            'responsable': responsable_id,
        }
    }
    return render(request, 'reportes/reporte_activos.html', context)


@login_required
@permission_required('reportes.view_reporte')
def reporte_mantenimientos(request):
    """Genera reporte de mantenimientos."""
    # Filtros de fecha
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tipo = request.GET.get('tipo')
    estado = request.GET.get('estado')
    
    if not fecha_desde:
        fecha_desde = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_hasta:
        fecha_hasta = timezone.now().strftime('%Y-%m-%d')
    
    queryset = Mantenimiento.objects.filter(
        fecha__gte=fecha_desde,
        fecha__lte=fecha_hasta
    ).select_related('activo', 'solicitado_por', 'realizado_por')
    
    if tipo:
        queryset = queryset.filter(tipo=tipo)
    if estado:
        queryset = queryset.filter(estado=estado)
    
    # Estadísticas
    stats = {
        'total': queryset.count(),
        'por_tipo': queryset.values('tipo').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad'),
        'por_estado': queryset.values('estado').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad'),
        'costo_total': queryset.aggregate(
            total=Sum('costo_total')
        )['total'] or 0,
        'tiempo_total': queryset.aggregate(
            total=Sum('tiempo_real')
        )['total'] or 0,
    }
    
    context = {
        'mantenimientos': queryset[:100],
        'stats': stats,
        'tipos': Mantenimiento.TIPO_MANTENIMIENTO,
        'estados': Mantenimiento.ESTADO_MANTENIMIENTO,
        'filtros': {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'tipo': tipo,
            'estado': estado,
        }
    }
    return render(request, 'reportes/reporte_mantenimientos.html', context)


@login_required
@permission_required('reportes.view_reporte')
def reporte_equipos(request):
    """Genera reporte de equipos funcionales."""
    queryset = EquipoFuncional.objects.select_related(
        'area__nivel__planta', 'responsable'
    ).annotate(activos_count=Count('activos'))
    
    # Estadísticas
    stats = {
        'total': queryset.count(),
        'por_estado': queryset.values('estado_operativo').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad'),
        'por_criticidad': queryset.values('criticidad').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad'),
        'por_planta': queryset.values('area__nivel__planta__nombre').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad'),
    }
    
    context = {
        'equipos': queryset[:100],
        'stats': stats,
    }
    return render(request, 'reportes/reporte_equipos.html', context)


# =============================================================================
# EXPORTACIONES A EXCEL
# =============================================================================

def crear_estilos_excel():
    """Crea estilos comunes para archivos Excel."""
    estilos = {
        'header': Font(bold=True, color='FFFFFF', size=11),
        'header_fill': PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'center': Alignment(horizontal='center', vertical='center'),
        'left': Alignment(horizontal='left', vertical='center'),
    }
    return estilos


@login_required
@permission_required('reportes.view_reporte')
def exportar_activos_excel(request):
    """Exporta activos a Excel."""
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activos"
    
    estilos = crear_estilos_excel()
    
    # Headers
    headers = [
        'Código', 'Nombre', 'Tipo', 'Estado', 'Propiedad',
        'Planta', 'Nivel', 'Área', 'Responsable',
        'Marca', 'Modelo', 'No. Serie',
        'Fecha Adquisición', 'Costo', 'Valor Actual'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = estilos['header']
        cell.fill = estilos['header_fill']
        cell.alignment = estilos['center']
        cell.border = estilos['border']
    
    # Datos
    activos = Activo.objects.select_related(
        'tipo', 'estado', 'tipo_propiedad', 'responsable', 'ubicacion'
    )
    
    for row, activo in enumerate(activos, 2):
        ws.cell(row=row, column=1, value=activo.codigo)
        ws.cell(row=row, column=2, value=activo.nombre)
        ws.cell(row=row, column=3, value=activo.tipo.nombre if activo.tipo else '')
        ws.cell(row=row, column=4, value=activo.estado.nombre if activo.estado else '')
        ws.cell(row=row, column=5, value=activo.tipo_propiedad.nombre if activo.tipo_propiedad else '')
        ws.cell(row=row, column=6, value=activo.ubicacion.planta.nombre if activo.ubicacion else '')
        ws.cell(row=row, column=7, value=activo.ubicacion.nivel.nombre if activo.ubicacion else '')
        ws.cell(row=row, column=8, value=activo.ubicacion.area.nombre if activo.ubicacion else '')
        ws.cell(row=row, column=9, value=str(activo.responsable) if activo.responsable else '')
        ws.cell(row=row, column=10, value=activo.marca)
        ws.cell(row=row, column=11, value=activo.modelo)
        ws.cell(row=row, column=12, value=activo.numero_serie)
        ws.cell(row=row, column=13, value=activo.fecha_adquisicion)
        ws.cell(row=row, column=14, value=float(activo.costo_adquisicion) if activo.costo_adquisicion else 0)
        ws.cell(row=row, column=15, value=float(activo.valor_actual) if activo.valor_actual else 0)
    
    # Ajustar anchos
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=activos.xlsx'
    wb.save(response)
    return response


@login_required
@permission_required('reportes.view_reporte')
def exportar_mantenimientos_excel(request):
    """Exporta mantenimientos a Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"
    
    estilos = crear_estilos_excel()
    
    headers = [
        'Código', 'Activo', 'Tipo', 'Fecha', 'Estado', 'Prioridad',
        'Descripción', 'Realizado por', 'Costo Total', 'Tiempo Real'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = estilos['header']
        cell.fill = estilos['header_fill']
        cell.alignment = estilos['center']
        cell.border = estilos['border']
    
    mantenimientos = Mantenimiento.objects.select_related(
        'activo', 'realizado_por'
    )
    
    for row, mtto in enumerate(mantenimientos, 2):
        ws.cell(row=row, column=1, value=mtto.codigo)
        ws.cell(row=row, column=2, value=str(mtto.activo))
        ws.cell(row=row, column=3, value=mtto.get_tipo_display())
        ws.cell(row=row, column=4, value=mtto.fecha)
        ws.cell(row=row, column=5, value=mtto.get_estado_display())
        ws.cell(row=row, column=6, value=mtto.get_prioridad_display())
        ws.cell(row=row, column=7, value=mtto.descripcion)
        ws.cell(row=row, column=8, value=str(mtto.realizado_por) if mtto.realizado_por else '')
        ws.cell(row=row, column=9, value=float(mtto.costo_total) if mtto.costo_total else 0)
        ws.cell(row=row, column=10, value=float(mtto.tiempo_real) if mtto.tiempo_real else 0)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=mantenimientos.xlsx'
    wb.save(response)
    return response


# =============================================================================
# DASHBOARD API
# =============================================================================

@login_required
def dashboard_data(request):
    """Retorna datos para el dashboard en formato JSON."""
    from django.http import JsonResponse
    
    hoy = timezone.now().date()
    hace_30_dias = hoy - timedelta(days=30)
    
    data = {
        'activos': {
            'total': Activo.objects.count(),
            'operativos': Activo.objects.filter(estado__operativo=True).count(),
            'fuera_servicio': Activo.objects.filter(estado__operativo=False).count(),
        },
        'mantenimientos': {
            'pendientes': Mantenimiento.objects.filter(estado='PENDIENTE').count(),
            'completados_mes': Mantenimiento.objects.filter(
                estado='COMPLETADO',
                fecha_realizacion__month=hoy.month,
                fecha_realizacion__year=hoy.year
            ).count(),
            'vencidos': Mantenimiento.objects.filter(
                estado='PENDIENTE',
                fecha__lt=hoy
            ).count(),
        },
        'equipos': {
            'total': EquipoFuncional.objects.count(),
            'operativos': EquipoFuncional.objects.filter(estado_operativo='OPERATIVO').count(),
        },
        'activos_por_tipo': list(Activo.objects.values('tipo__nombre').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad')[:5]),
        'mantenimientos_por_mes': list(Mantenimiento.objects.filter(
            fecha__gte=hace_30_dias
        ).values('fecha__month').annotate(
            cantidad=Count('id')
        ).order_by('fecha__month')),
    }
    
    return JsonResponse(data)


# =============================================================================
# EXPORTACIÓN A PDF
# =============================================================================

import io
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable, Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer,
    Table, TableStyle,
)

# Paleta de colores
C_PRIMARY   = colors.HexColor("#1e3a5f")
C_ACCENT    = colors.HexColor("#0ea5e9")
C_ACCENT2   = colors.HexColor("#06b6d4")
C_GREEN     = colors.HexColor("#22c55e")
C_ORANGE    = colors.HexColor("#f59e0b")
C_SURFACE   = colors.HexColor("#f1f5f9")
C_BORDER    = colors.HexColor("#e2e8f0")
C_TEXT      = colors.HexColor("#1e293b")
C_MUTED     = colors.HexColor("#64748b")
C_WHITE     = colors.white
C_BLACK     = colors.black

PRIO_COLORS = {
    "BAJA":    colors.HexColor("#4ade80"),
    "MEDIA":   colors.HexColor("#60a5fa"),
    "ALTA":    colors.HexColor("#fb923c"),
    "URGENTE": colors.HexColor("#f87171"),
}

def _styles():
    base = getSampleStyleSheet()
    def ps(name, **kw):
        return ParagraphStyle(name, **kw)
    return {
        "title": ps("title", fontName="Helvetica-Bold", fontSize=16,
                    textColor=C_PRIMARY, spaceAfter=2),
        "subtitle": ps("subtitle", fontName="Helvetica", fontSize=9,
                       textColor=C_MUTED, spaceAfter=0),
        "section": ps("section", fontName="Helvetica-Bold", fontSize=9,
                      textColor=C_PRIMARY, spaceBefore=10, spaceAfter=4,
                      textTransform="uppercase", letterSpacing=0.8),
        "normal": ps("normal", fontName="Helvetica", fontSize=8, textColor=C_TEXT),
        "small": ps("small", fontName="Helvetica", fontSize=7, textColor=C_MUTED),
        "th": ps("th", fontName="Helvetica-Bold", fontSize=7.5,
                 textColor=C_WHITE, alignment=TA_LEFT),
        "td": ps("td", fontName="Helvetica", fontSize=7.5,
                 textColor=C_TEXT, alignment=TA_LEFT),
        "td_code": ps("td_code", fontName="Helvetica-Bold", fontSize=7.5,
                      textColor=C_ACCENT2, alignment=TA_LEFT),
        "td_num": ps("td_num", fontName="Helvetica-Bold", fontSize=7.5,
                     textColor=C_GREEN, alignment=TA_RIGHT),
        "sign_name": ps("sign_name", fontName="Helvetica-Bold", fontSize=8.5,
                        textColor=C_TEXT, alignment=TA_CENTER),
        "sign_role": ps("sign_role", fontName="Helvetica", fontSize=7.5,
                        textColor=C_MUTED, alignment=TA_CENTER),
        "footer": ps("footer", fontName="Helvetica", fontSize=7,
                     textColor=C_MUTED, alignment=TA_CENTER),
        "stat_num": ps("stat_num", fontName="Helvetica-Bold", fontSize=18,
                       textColor=C_PRIMARY, alignment=TA_CENTER, leading=20),
        "stat_lbl": ps("stat_lbl", fontName="Helvetica", fontSize=7,
                       textColor=C_MUTED, alignment=TA_CENTER),
    }

INSTITUTION = "SIGAP · Sistema de Gestión de Activos y Proyectos"
REPORT_TITLE = "REPORTE DE MANTENIMIENTOS"

def _header_footer(canvas, doc):
    canvas.saveState()
    w, h = A4
    canvas.setFillColor(C_PRIMARY)
    canvas.rect(0, h - 2*cm, w, 2*cm, fill=1, stroke=0)
    canvas.setFont("Helvetica-Bold", 11)
    canvas.setFillColor(C_WHITE)
    canvas.drawString(1.5*cm, h - 1.1*cm, REPORT_TITLE)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#94a3b8"))
    canvas.drawString(1.5*cm, h - 1.55*cm, INSTITUTION)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#bae6fd"))
    canvas.drawRightString(w - 1.5*cm, h - 1.1*cm,
                           f"Emitido: {date.today().strftime('%d/%m/%Y')}")
    canvas.setStrokeColor(C_ACCENT2)
    canvas.setLineWidth(1.5)
    canvas.line(0, h - 2*cm - 1, w, h - 2*cm - 1)
    canvas.setStrokeColor(C_BORDER)
    canvas.setLineWidth(0.5)
    canvas.line(1.5*cm, 1.5*cm, w - 1.5*cm, 1.5*cm)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(C_MUTED)
    canvas.drawString(1.5*cm, 1.1*cm,
                      "Documento generado automáticamente — No requiere sello húmedo")
    canvas.drawRightString(w - 1.5*cm, 1.1*cm,
                           f"Página {doc.page}")
    canvas.restoreState()

def _stat_block(stats, styles):
    items = [
        (str(stats.get("total", 0)),           "Total registros",  C_ACCENT),
        (str(len(stats.get("por_tipo", []))),   "Tipos distintos",  C_ACCENT2),
        (f"{stats.get('costo_total', 0):,.0f}", "Costo total (BOB)",C_GREEN),
        (f"{stats.get('tiempo_total', 0):.1f}h","Tiempo total",     C_ORANGE),
    ]
    cells = []
    for num, lbl, color in items:
        p_num = Paragraph(f'<font color="{color.hexval()}">{num}</font>',
                          styles["stat_num"])
        p_lbl = Paragraph(lbl, styles["stat_lbl"])
        cells.append([p_num, p_lbl])
    data_row = [[cells[i][0] for i in range(4)],
                [cells[i][1] for i in range(4)]]
    col_w = (A4[0] - 3*cm) / 4
    tbl = Table(data_row, colWidths=[col_w]*4, rowHeights=[1*cm, 0.5*cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), C_SURFACE),
        ("BOX",        (0,0), (-1,-1), 0.5, C_BORDER),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, C_BORDER),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("ROUNDEDCORNERS", [4]),
    ]))
    return tbl

def _main_table(mantenimientos, styles):
    header = ["Código", "Activo", "Tipo", "Fecha", "Estado", "Prioridad",
              "Costo (BOB)", "Tiempo"]
    th_cells = [Paragraph(h, styles["th"]) for h in header]
    rows = [th_cells]
    for m in mantenimientos:
        prio_key = getattr(m, "prioridad", "BAJA")
        prio_color = PRIO_COLORS.get(prio_key, C_MUTED)
        prio_label = m.get_prioridad_display() if hasattr(m, "get_prioridad_display") else prio_key
        prio_cell = Paragraph(
            f'<font color="{prio_color.hexval()}">● </font>'
            f'<font color="{C_MUTED.hexval()}">{prio_label}</font>',
            styles["td"])
        costo = m.costo_total
        costo_str = f"{costo:,.0f}" if costo else "—"
        row = [
            Paragraph(m.codigo or "—",            styles["td_code"]),
            Paragraph(m.activo.nombre,             styles["td"]),
            Paragraph(m.get_tipo_display(),        styles["td"]),
            Paragraph(m.fecha.strftime("%d/%m/%Y"),styles["td"]),
            Paragraph(m.get_estado_display(),      styles["td"]),
            prio_cell,
            Paragraph(costo_str,                   styles["td_num"]),
            Paragraph(f"{m.tiempo_real}h" if m.tiempo_real else "—", styles["td"]),
        ]
        rows.append(row)
    col_widths = [2.2*cm, 4.5*cm, 2.4*cm, 2.3*cm, 2.4*cm, 2.3*cm, 2.2*cm, 1.8*cm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  C_PRIMARY),
        ("TEXTCOLOR",     (0,0), (-1,0),  C_WHITE),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  7.5),
        ("TOPPADDING",    (0,0), (-1,0),  7),
        ("BOTTOMPADDING", (0,0), (-1,0),  7),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_WHITE, C_SURFACE]),
        ("FONTSIZE",      (0,1), (-1,-1), 7.5),
        ("TOPPADDING",    (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ("BOX",           (0,0), (-1,-1), 0.5, C_BORDER),
        ("INNERGRID",     (0,0), (-1,-1), 0.3, C_BORDER),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("LINEBELOW",     (0,0), (-1,0),  1.5, C_ACCENT2),
    ]))
    return tbl

def _dist_table(items, key_field, title, bar_color, styles):
    if not items:
        return Paragraph("Sin datos", styles["small"])
    max_val = items[0]["cantidad"] if items else 1
    rows = []
    for item in items:
        pct = int((item["cantidad"] / max_val) * 60)
        bar = Table([[""]], colWidths=[pct * 0.8 + 4], rowHeights=[6])
        bar.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), bar_color),
            ("ROUNDEDCORNERS", [3]),
            ("TOPPADDING",    (0,0),(-1,-1),0),
            ("BOTTOMPADDING", (0,0),(-1,-1),0),
            ("LEFTPADDING",   (0,0),(-1,-1),0),
            ("RIGHTPADDING",  (0,0),(-1,-1),0),
        ]))
        rows.append([
            Paragraph(str(item[key_field]), styles["small"]),
            bar,
            Paragraph(str(item["cantidad"]),
                      ParagraphStyle("dc", fontName="Helvetica-Bold",
                                     fontSize=7.5, textColor=C_TEXT,
                                     alignment=TA_RIGHT)),
        ])
    tbl = Table(rows, colWidths=[3.8*cm, 2*cm, 0.9*cm])
    tbl.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1),3),
        ("BOTTOMPADDING", (0,0),(-1,-1),3),
        ("LEFTPADDING",   (0,0),(-1,-1),4),
        ("RIGHTPADDING",  (0,0),(-1,-1),4),
        ("LINEBELOW",     (0,0),(-1,-2), 0.3, C_BORDER),
    ]))
    return tbl

def _firma_block(firmantes, styles):
    LINE_W = 4 * cm
    col_data = []
    for f in firmantes:
        ci_str = f"C.I. {f['ci']}" if f.get("ci") else ""
        cell = [
            Table([[""]], colWidths=[LINE_W], rowHeights=[0.3],
                  style=[("LINEABOVE", (0,0),(-1,-1), 0.8, C_TEXT),
                         ("TOPPADDING",(0,0),(-1,-1),0),
                         ("BOTTOMPADDING",(0,0),(-1,-1),0)]),
            Spacer(1, 3),
            Paragraph(f["nombre"], styles["sign_name"]),
            Paragraph(f["cargo"],  styles["sign_role"]),
        ]
        if ci_str:
            cell.append(Paragraph(ci_str, styles["sign_role"]))
        col_data.append(cell)
    n = len(firmantes)
    page_w = A4[0] - 3*cm
    col_w  = page_w / n
    row = [[item for item in col] for col in col_data]
    tbl = Table([row], colWidths=[col_w]*n)
    tbl.setStyle(TableStyle([
        ("VALIGN",  (0,0),(-1,-1),"TOP"),
        ("ALIGN",   (0,0),(-1,-1),"CENTER"),
        ("TOPPADDING",    (0,0),(-1,-1),0),
        ("BOTTOMPADDING", (0,0),(-1,-1),0),
        ("LEFTPADDING",   (0,0),(-1,-1),4),
        ("RIGHTPADDING",  (0,0),(-1,-1),4),
    ]))
    return tbl

def _build_pdf(buffer, mantenimientos, stats, filtros, firmantes):
    styles = _styles()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=2.5*cm,
        bottomMargin=2.2*cm,
        title="Reporte de Mantenimientos",
        author=INSTITUTION,
    )
    story = []
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Reporte de Mantenimientos", styles["title"]))
    story.append(Paragraph(
        f"Período: {filtros['fecha_desde']} al {filtros['fecha_hasta']}"
        + (f"  ·  Tipo: {filtros['tipo']}" if filtros.get("tipo") else "")
        + (f"  ·  Estado: {filtros['estado']}" if filtros.get("estado") else ""),
        styles["subtitle"]))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=C_BORDER, spaceAfter=10))
    story.append(Paragraph("Resumen ejecutivo", styles["section"]))
    story.append(_stat_block(stats, styles))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("Distribución", styles["section"]))
    def dist_header(title, styles):
        return Paragraph(title.upper(), ParagraphStyle(
            "dh", fontName="Helvetica-Bold", fontSize=7,
            textColor=C_MUTED, spaceAfter=4, letterSpacing=0.7))
    half = (A4[0] - 3*cm) / 2
    dist_tipo   = _dist_table(stats["por_tipo"],   "tipo",   "Por tipo",   C_ACCENT,  styles)
    dist_estado = _dist_table(stats["por_estado"], "estado", "Por estado", C_ACCENT2, styles)
    dist_block = Table(
        [[dist_header("Por tipo", styles), dist_header("Por estado", styles)],
         [dist_tipo,                       dist_estado]],
        colWidths=[half, half]
    )
    dist_block.setStyle(TableStyle([
        ("VALIGN",      (0,0),(-1,-1),"TOP"),
        ("TOPPADDING",  (0,0),(-1,-1),0),
        ("LEFTPADDING", (0,0),(-1,-1),0),
        ("RIGHTPADDING",(0,0),(-1,-1),8),
    ]))
    story.append(dist_block)
    story.append(Spacer(1, 0.6*cm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=C_BORDER, spaceBefore=2, spaceAfter=8))
    story.append(Paragraph(
        f"Detalle de mantenimientos ({stats['total']} registros)",
        styles["section"]))
    if mantenimientos:
        story.append(_main_table(mantenimientos, styles))
    else:
        story.append(Paragraph("Sin mantenimientos para el período seleccionado.",
                                styles["small"]))
    story.append(Spacer(1, 1.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=C_BORDER, spaceAfter=8))
    story.append(Paragraph("Conformidad y Firmas", styles["section"]))
    story.append(Spacer(1, 0.8*cm))
    story.append(_firma_block(firmantes, styles))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Fecha de emisión: {date.today().strftime('%d de %B de %Y')}  ·  "
        "Documento válido sin sello húmedo conforme normativa interna.",
        styles["footer"]))
    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

@login_required
@permission_required('reportes.view_reporte')
def exportar_mantenimientos_pdf(request):
    """Exporta reporte de mantenimientos a PDF."""
    fecha_desde = request.GET.get("fecha_desde") or "2024-01-01"
    fecha_hasta = request.GET.get("fecha_hasta") or date.today().isoformat()
    tipo        = request.GET.get("tipo", "")
    estado      = request.GET.get("estado", "")

    # Usamos el modelo importado correctamente
    from apps.mantenimiento.models import Mantenimiento

    qs = (Mantenimiento.objects
          .select_related("activo")
          .filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta))
    if tipo:
        qs = qs.filter(tipo=tipo)
    if estado:
        qs = qs.filter(estado=estado)

    mantenimientos = qs.order_by("-fecha")[:200]

    from django.db.models import Sum, Count
    agg = qs.aggregate(
        costo_total=Sum("costo_total"),
        tiempo_total=Sum("tiempo_real"),
    )
    por_tipo   = (qs.values("tipo")
                    .annotate(cantidad=Count("id"))
                    .order_by("-cantidad"))
    por_estado = (qs.values("estado")
                    .annotate(cantidad=Count("id"))
                    .order_by("-cantidad"))

    stats = {
        "total":       qs.count(),
        "costo_total": agg["costo_total"] or 0,
        "tiempo_total":agg["tiempo_total"] or 0,
        "por_tipo":    list(por_tipo.values("tipo", "cantidad")),
        "por_estado":  list(por_estado.values("estado", "cantidad")),
    }

    filtros = {
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "tipo":  tipo,
        "estado":estado,
    }

    firmantes = [
        {"nombre": "Ing. Miguel Gutierrez Magne",  "cargo": "Responsable PPPI",  "ci": ""},
        {"nombre": " ",  "cargo": "Responsable de Activos"},
        {"nombre": " ","cargo": "Director Técnico",        "ci": ""},
    ]

    buffer = io.BytesIO()
    _build_pdf(buffer, mantenimientos, stats, filtros, firmantes)
    buffer.seek(0)

    resp = HttpResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = (
        f'attachment; filename="reporte_mantenimientos_{date.today()}.pdf"'
    )
    return resp


    # =============================================================================
# EXPORTACIÓN A PDF - ACTIVOS
# =============================================================================

@login_required
@permission_required('reportes.view_reporte')
def exportar_activos_pdf(request):
    """Exporta reporte de activos a PDF."""
    from apps.activos.models import Activo
    from django.db.models import Sum, Count

    # Filtros igual que en reporte_activos
    planta_id = request.GET.get('planta')
    tipo_id = request.GET.get('tipo')
    estado_id = request.GET.get('estado')
    responsable_id = request.GET.get('responsable')

    queryset = Activo.objects.select_related('tipo', 'estado', 'responsable', 'ubicacion')

    if planta_id:
        queryset = queryset.filter(ubicacion__planta_id=planta_id)
    if tipo_id:
        queryset = queryset.filter(tipo_id=tipo_id)
    if estado_id:
        queryset = queryset.filter(estado_id=estado_id)
    if responsable_id:
        queryset = queryset.filter(responsable_id=responsable_id)

    activos = queryset[:200]  # límite para no sobrecargar el PDF

    # Estadísticas
    stats = {
        'total': queryset.count(),
        'por_tipo': queryset.values('tipo__nombre').annotate(cantidad=Count('id')).order_by('-cantidad'),
        'por_estado': queryset.values('estado__nombre').annotate(cantidad=Count('id')).order_by('-cantidad'),
        'por_planta': queryset.values('ubicacion__planta__nombre').annotate(cantidad=Count('id')).order_by('-cantidad'),
        'valor_total': queryset.aggregate(total=Sum('valor_actual'))['total'] or 0,
    }

    filtros = {
        'planta': planta_id,
        'tipo': tipo_id,
        'estado': estado_id,
        'responsable': responsable_id,
    }

    firmantes = [
        {"nombre": "Ing. Miguel Gutierrez Magne", "cargo": "Responsable PPPI", "ci": ""},
        {"nombre": "", "cargo": "Responsable de Activos"},
        {"nombre": "", "cargo": "Director Técnico", "ci": ""},
    ]

    buffer = io.BytesIO()
    _build_pdf_activos(buffer, activos, stats, filtros, firmantes)
    buffer.seek(0)

    resp = HttpResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="reporte_activos_{date.today()}.pdf"'
    return resp


def _build_pdf_activos(buffer, activos, stats, filtros, firmantes):
    """Construye el PDF para activos."""
    styles = _styles()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=2.5*cm,
        bottomMargin=2.2*cm,
        title="Reporte de Activos",
        author=INSTITUTION,
    )
    story = []

    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Reporte de Activos", styles["title"]))

    # Filtros textuales
    filtro_texto = []
    if filtros.get('planta'):
        planta_nombre = Planta.objects.filter(id=filtros['planta']).values_list('nombre', flat=True).first()
        if planta_nombre:
            filtro_texto.append(f"Planta: {planta_nombre}")
    if filtros.get('tipo'):
        tipo_nombre = TipoActivo.objects.filter(id=filtros['tipo']).values_list('nombre', flat=True).first()
        if tipo_nombre:
            filtro_texto.append(f"Tipo: {tipo_nombre}")
    if filtros.get('estado'):
        estado_nombre = Estado.objects.filter(id=filtros['estado']).values_list('nombre', flat=True).first()
        if estado_nombre:
            filtro_texto.append(f"Estado: {estado_nombre}")
    if filtros.get('responsable'):
        responsable_nombre = Responsable.objects.filter(id=filtros['responsable']).values_list('nombre', flat=True).first()
        if responsable_nombre:
            filtro_texto.append(f"Responsable: {responsable_nombre}")

    story.append(Paragraph(
        " · ".join(filtro_texto) if filtro_texto else "Todos los activos",
        styles["subtitle"]))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=10))

    # Estadísticas
    story.append(Paragraph("Resumen ejecutivo", styles["section"]))
    stats_items = [
        (str(stats['total']), "Total activos", C_ACCENT),
        (str(len(stats['por_tipo'])), "Tipos distintos", C_ACCENT2),
        (str(len(stats['por_planta'])), "Plantas", C_ORANGE),
        (f"{stats['valor_total']:,.0f}", "Valor total (BOB)", C_GREEN),
    ]
    story.append(_stat_block_custom(stats_items, styles))
    story.append(Spacer(1, 0.5*cm))

    # Distribuciones
    story.append(Paragraph("Distribución", styles["section"]))
    def dist_header(title):
        return Paragraph(title.upper(), ParagraphStyle(
            "dh", fontName="Helvetica-Bold", fontSize=7,
            textColor=C_MUTED, spaceAfter=4, letterSpacing=0.7))

    half = (A4[0] - 3*cm) / 2
    dist_tipo = _dist_table(stats['por_tipo'], 'tipo__nombre', "Por tipo", C_ACCENT, styles)
    dist_estado = _dist_table(stats['por_estado'], 'estado__nombre', "Por estado", C_ACCENT2, styles)
    dist_planta = _dist_table(stats['por_planta'], 'ubicacion__planta__nombre', "Por planta", colors.HexColor("#8b5cf6"), styles)

    dist_block = Table(
        [[dist_header("Por tipo"), dist_header("Por estado")],
         [dist_tipo, dist_estado]],
        colWidths=[half, half]
    )
    dist_block.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
    story.append(dist_block)
    story.append(Spacer(1, 0.6*cm))

    # Tabla principal
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceBefore=2, spaceAfter=8))
    story.append(Paragraph(f"Detalle de activos ({stats['total']} registros)", styles["section"]))
    if activos:
        story.append(_main_table_activos(activos, styles))
    else:
        story.append(Paragraph("Sin activos para los filtros seleccionados.", styles["small"]))

    # Firmas
    story.append(Spacer(1, 1.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=8))
    story.append(Paragraph("Conformidad y Firmas", styles["section"]))
    story.append(Spacer(1, 0.8*cm))
    story.append(_firma_block(firmantes, styles))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Fecha de emisión: {date.today().strftime('%d de %B de %Y')}  ·  "
        "Documento válido sin sello húmedo conforme normativa interna.",
        styles["footer"]))

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)


def _stat_block_custom(items, styles):
    """items: lista de (num, lbl, color)"""
    cells = []
    for num, lbl, color in items:
        p_num = Paragraph(f'<font color="{color.hexval()}">{num}</font>', styles["stat_num"])
        p_lbl = Paragraph(lbl, styles["stat_lbl"])
        cells.append([p_num, p_lbl])
    data_row = [[cells[i][0] for i in range(4)],
                [cells[i][1] for i in range(4)]]
    col_w = (A4[0] - 3*cm) / 4
    tbl = Table(data_row, colWidths=[col_w]*4, rowHeights=[1*cm, 0.5*cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), C_SURFACE),
        ("BOX", (0,0), (-1,-1), 0.5, C_BORDER),
        ("INNERGRID", (0,0), (-1,-1), 0.3, C_BORDER),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("ROUNDEDCORNERS", [4]),
    ]))
    return tbl


def _main_table_activos(activos, styles):
    header = ["Código", "Nombre", "Tipo", "Estado", "Planta", "Responsable", "Valor (BOB)"]
    th_cells = [Paragraph(h, styles["th"]) for h in header]
    rows = [th_cells]

    for a in activos:
        valor = a.valor_actual
        valor_str = f"{valor:,.0f}" if valor else "—"
        row = [
            Paragraph(a.codigo or "—", styles["td_code"]),
            Paragraph(a.nombre, styles["td"]),
            Paragraph(a.tipo.nombre if a.tipo else "—", styles["td"]),
            Paragraph(a.estado.nombre if a.estado else "—", styles["td"]),
            Paragraph(a.ubicacion.planta.nombre if a.ubicacion and a.ubicacion.planta else "—", styles["td"]),
            Paragraph(str(a.responsable) if a.responsable else "—", styles["td"]),
            Paragraph(valor_str, styles["td_num"]),
        ]
        rows.append(row)

    col_widths = [2.2*cm, 4.5*cm, 2.4*cm, 2.3*cm, 2.4*cm, 3*cm, 2.2*cm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), C_PRIMARY),
        ("TEXTCOLOR", (0,0), (-1,0), C_WHITE),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 7.5),
        ("TOPPADDING", (0,0), (-1,0), 7),
        ("BOTTOMPADDING", (0,0), (-1,0), 7),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [C_WHITE, C_SURFACE]),
        ("FONTSIZE", (0,1), (-1,-1), 7.5),
        ("TOPPADDING", (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("BOX", (0,0), (-1,-1), 0.5, C_BORDER),
        ("INNERGRID", (0,0), (-1,-1), 0.3, C_BORDER),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LINEBELOW", (0,0), (-1,0), 1.5, C_ACCENT2),
    ]))
    return tbl



    # =============================================================================
# EXPORTACIÓN A PDF - EQUIPOS
# =============================================================================

@login_required
@permission_required('reportes.view_reporte')
def exportar_equipos_pdf(request):
    """Exporta reporte de equipos funcionales a PDF."""
    from apps.activos.models import EquipoFuncional
    from django.db.models import Count

    queryset = EquipoFuncional.objects.select_related(
        'area__nivel__planta', 'responsable'
    ).annotate(activos_count=Count('activos'))

    equipos = queryset[:200]

    stats = {
        'total': queryset.count(),
        'por_estado': list(queryset.values('estado_operativo')
                           .annotate(cantidad=Count('id'))
                           .order_by('-cantidad')),
        'por_criticidad': list(queryset.values('criticidad')
                               .annotate(cantidad=Count('id'))
                               .order_by('-cantidad')),
        'por_planta': list(queryset.values('area__nivel__planta__nombre')
                           .annotate(cantidad=Count('id'))
                           .order_by('-cantidad')),
    }

    firmantes = [
        {"nombre": "Ing. Juan Pérez Flores", "cargo": "Jefe de Mantenimiento", "ci": "4512389"},
        {"nombre": "Lic. Ana Mamani Quispe", "cargo": "Responsable de Activos"},
        {"nombre": "Dir. Carlos Rojas Vargas", "cargo": "Director Técnico", "ci": "7823140"},
    ]

    buffer = io.BytesIO()
    _build_pdf_equipos(buffer, equipos, stats, firmantes)
    buffer.seek(0)

    resp = HttpResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="reporte_equipos_{date.today()}.pdf"'
    return resp


def _build_pdf_equipos(buffer, equipos, stats, firmantes):
    """Construye el PDF para equipos funcionales."""
    styles = _styles()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=2.5*cm,
        bottomMargin=2.2*cm,
        title="Reporte de Equipos Funcionales",
        author=INSTITUTION,
    )
    story = []

    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Reporte de Equipos Funcionales", styles["title"]))
    story.append(Paragraph("Todos los equipos registrados", styles["subtitle"]))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=10))

    # Estadísticas
    story.append(Paragraph("Resumen ejecutivo", styles["section"]))
    total_operativos = next((item['cantidad'] for item in stats['por_estado'] if item['estado_operativo'] == 'OPERATIVO'), 0)
    total_fuera = next((item['cantidad'] for item in stats['por_estado'] if item['estado_operativo'] == 'FUERA_SERVICIO'), 0)
    stats_items = [
        (str(stats['total']), "Total equipos", C_ACCENT),
        (str(total_operativos), "Operativos", C_GREEN),
        (str(len(stats['por_criticidad'])), "Niveles criticidad", C_ORANGE),
        (str(total_fuera), "Fuera de servicio", colors.HexColor("#f87171")),
    ]
    story.append(_stat_block_custom(stats_items, styles))
    story.append(Spacer(1, 0.5*cm))

    # Distribuciones
    story.append(Paragraph("Distribución", styles["section"]))
    def dist_header(title):
        return Paragraph(title.upper(), ParagraphStyle(
            "dh", fontName="Helvetica-Bold", fontSize=7,
            textColor=C_MUTED, spaceAfter=4, letterSpacing=0.7))

    half = (A4[0] - 3*cm) / 2

    # Tabla para estado operativo (con puntos de color)
    def _dist_table_estado(items):
        if not items:
            return Paragraph("Sin datos", styles["small"])
        max_val = items[0]['cantidad'] if items else 1
        rows = []
        for item in items:
            estado = item['estado_operativo']
            color = {
                'OPERATIVO': '#22c55e',
                'FUERA_SERVICIO': '#ef4444',
                'MANTENIMIENTO': '#f59e0b',
                'EN_REPARACION': '#3b82f6'
            }.get(estado, C_MUTED.hexval())
            bar = Table([[""]], colWidths=[int((item['cantidad']/max_val)*60)*0.8+4], rowHeights=[6])
            bar.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.HexColor(color)),
                                     ("ROUNDEDCORNERS", [3]),
                                     ("TOPPADDING",(0,0),(-1,-1),0),
                                     ("BOTTOMPADDING",(0,0),(-1,-1),0),
                                     ("LEFTPADDING",(0,0),(-1,-1),0),
                                     ("RIGHTPADDING",(0,0),(-1,-1),0)]))
            rows.append([
                Paragraph(f'<font color="{color}">● </font> {estado}', styles["small"]),
                bar,
                Paragraph(str(item['cantidad']),
                          ParagraphStyle("dc", fontName="Helvetica-Bold",
                                         fontSize=7.5, textColor=C_TEXT,
                                         alignment=TA_RIGHT)),
            ])
        tbl = Table(rows, colWidths=[3.8*cm, 2*cm, 0.9*cm])
        tbl.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                                 ("TOPPADDING", (0,0), (-1,-1), 3),
                                 ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                                 ("LEFTPADDING", (0,0), (-1,-1), 4),
                                 ("RIGHTPADDING", (0,0), (-1,-1), 4),
                                 ("LINEBELOW", (0,0), (-1,-2), 0.3, C_BORDER)]))
        return tbl

    # Tabla para criticidad (con colores)
    def _dist_table_criticidad(items):
        if not items:
            return Paragraph("Sin datos", styles["small"])
        max_val = items[0]['cantidad'] if items else 1
        rows = []
        for item in items:
            crit = item['criticidad']
            color = {'ALTA': '#ef4444', 'MEDIA': '#f59e0b', 'BAJA': '#22c55e'}.get(crit, C_MUTED.hexval())
            bar = Table([[""]], colWidths=[int((item['cantidad']/max_val)*60)*0.8+4], rowHeights=[6])
            bar.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.HexColor(color)),
                                     ("ROUNDEDCORNERS", [3]),
                                     ("TOPPADDING",(0,0),(-1,-1),0),
                                     ("BOTTOMPADDING",(0,0),(-1,-1),0),
                                     ("LEFTPADDING",(0,0),(-1,-1),0),
                                     ("RIGHTPADDING",(0,0),(-1,-1),0)]))
            rows.append([
                Paragraph(f'<font color="{color}">{crit}</font>', styles["small"]),
                bar,
                Paragraph(str(item['cantidad']),
                          ParagraphStyle("dc", fontName="Helvetica-Bold",
                                         fontSize=7.5, textColor=C_TEXT,
                                         alignment=TA_RIGHT)),
            ])
        tbl = Table(rows, colWidths=[3.8*cm, 2*cm, 0.9*cm])
        tbl.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                                 ("TOPPADDING", (0,0), (-1,-1), 3),
                                 ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                                 ("LEFTPADDING", (0,0), (-1,-1), 4),
                                 ("RIGHTPADDING", (0,0), (-1,-1), 4),
                                 ("LINEBELOW", (0,0), (-1,-2), 0.3, C_BORDER)]))
        return tbl

    # Por planta (reutilizamos _dist_table genérica)
    dist_planta = _dist_table(stats['por_planta'], 'area__nivel__planta__nombre', "Por planta", colors.HexColor("#8b5cf6"), styles)

    dist_estado = _dist_table_estado(stats['por_estado'])
    dist_criticidad = _dist_table_criticidad(stats['por_criticidad'])

    dist_block = Table(
        [[dist_header("Por estado"), dist_header("Por criticidad")],
         [dist_estado, dist_criticidad]],
        colWidths=[half, half]
    )
    dist_block.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
    story.append(dist_block)
    story.append(Spacer(1, 0.6*cm))

    # Tabla principal
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceBefore=2, spaceAfter=8))
    story.append(Paragraph(f"Detalle de equipos ({stats['total']} registros)", styles["section"]))
    if equipos:
        story.append(_main_table_equipos(equipos, styles))
    else:
        story.append(Paragraph("Sin equipos registrados.", styles["small"]))

    # Firmas
    story.append(Spacer(1, 1.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=8))
    story.append(Paragraph("Conformidad y Firmas", styles["section"]))
    story.append(Spacer(1, 0.8*cm))
    story.append(_firma_block(firmantes, styles))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Fecha de emisión: {date.today().strftime('%d de %B de %Y')}  ·  "
        "Documento válido sin sello húmedo conforme normativa interna.",
        styles["footer"]))

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)


def _main_table_equipos(equipos, styles):
    header = ["Código", "Nombre", "Planta", "Estado operativo", "Criticidad", "Responsable", "Activos"]
    th_cells = [Paragraph(h, styles["th"]) for h in header]
    rows = [th_cells]

    for eq in equipos:
        # Estado con punto
        estado = eq.estado_operativo
        color = {
            'OPERATIVO': '#22c55e',
            'FUERA_SERVICIO': '#ef4444',
            'MANTENIMIENTO': '#f59e0b',
            'EN_REPARACION': '#3b82f6'
        }.get(estado, C_MUTED.hexval())
        estado_display = eq.get_estado_operativo_display() or estado
        estado_cell = Paragraph(
            f'<font color="{color}">● </font> {estado_display}',
            styles["td"])

        # Criticidad con color
        criticidad = eq.criticidad
        crit_color = {'ALTA': '#ef4444', 'MEDIA': '#f59e0b', 'BAJA': '#22c55e'}.get(criticidad, C_MUTED.hexval())
        crit_display = eq.get_criticidad_display() or criticidad
        crit_cell = Paragraph(
            f'<font color="{crit_color}">{crit_display}</font>',
            styles["td"])

        row = [
            Paragraph(eq.codigo or "—", styles["td_code"]),
            Paragraph(eq.nombre, styles["td"]),
            Paragraph(eq.area.nivel.planta.nombre if eq.area and eq.area.nivel and eq.area.nivel.planta else "—", styles["td"]),
            estado_cell,
            crit_cell,
            Paragraph(str(eq.responsable) if eq.responsable else "—", styles["td"]),
            Paragraph(str(eq.activos_count), styles["td_num"]),
        ]
        rows.append(row)

    col_widths = [2.2*cm, 4.5*cm, 2.4*cm, 2.5*cm, 2.0*cm, 2.8*cm, 1.2*cm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), C_PRIMARY),
        ("TEXTCOLOR", (0,0), (-1,0), C_WHITE),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 7.5),
        ("TOPPADDING", (0,0), (-1,0), 7),
        ("BOTTOMPADDING", (0,0), (-1,0), 7),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [C_WHITE, C_SURFACE]),
        ("FONTSIZE", (0,1), (-1,-1), 7.5),
        ("TOPPADDING", (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("BOX", (0,0), (-1,-1), 0.5, C_BORDER),
        ("INNERGRID", (0,0), (-1,-1), 0.3, C_BORDER),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LINEBELOW", (0,0), (-1,0), 1.5, C_ACCENT2),
    ]))
    return tbl